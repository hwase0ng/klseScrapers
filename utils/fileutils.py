'''
Created on Dec 20, 2016

@author: hwase0ng
'''
import settings as S
from datetime import datetime
import csv
import glob
import mmap
import openpyxl
import os
import socket
import subprocess
import time
import fnmatch


def mapcount(filename):
    f = open(filename, "r+")
    buf = mmap.mmap(f.fileno(), 0)
    lines = 0
    readline = buf.readline
    while readline():
        lines += 1
    return lines


def execshell(shellcmd):
    import shlex
    cmd = shlex.split(shellcmd)
    # out = subprocess.Popen(cmd)
    (out, _) = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, shell=True).communicate()
    result = out.replace('\n', '')
    return result


def grepN(filename, searchstr):
    return grep(filename, '-n', searchstr)


def grep(filename, option, searchstr):
    out = subprocess.Popen(['grep', option, searchstr, filename],
                           stdout=subprocess.PIPE,
                           stderr=subprocess.STDOUT
                           ).communicate()[0]
    if "No such file" in out:
        return -1
    try:
        result = int(out.partition(b':')[0])
    except Exception:
        result = -1
    return result


def wc_line_count(filename):
    out = subprocess.Popen(['wc', '-l', filename],
                           stdout=subprocess.PIPE,
                           stderr=subprocess.STDOUT
                           ).communicate()[0]
    if "No such file" in out:
        return -1
    return int(out.partition(b' ')[0])


def tail(fl):
    line = subprocess.check_output(['tail', '-1', fl])
    return line


def tail2(fl, n=1, bs=1024):
    try:
        f = open(fl)
        f.seek(-1, 2)
        l = 1-f.read(1).count('\n')  # If file doesn't end in \n, count anyway
        B = f.tell()
        while n >= l and B > 0:
                block = min(bs, B)
                B -= block
                f.seek(B, 0)
                l += f.read(block).count('\n')
        f.seek(B, 0)
        l = min(l, n)  # discard first (incomplete) line if l > n
        lines = f.readlines()[-l:]
        f.close()
        return lines
    except:
        if S.DBG_ALL:
            print "tail has failed on ", fl
        return [""]


def getSystemIP():
    hostname = socket.gethostname()
    ip = socket.gethostbyname(hostname)
#   print type(ip), ip
    return ip


def concat2quotes(directory, target):
    with cd(directory):
        if S.DBG_ALL:
            print os.getcwd()
#       os.system("type *.csv >> quotes.txt")
        try:
            with open(S.DATA_DIR + S.SHORTLISTED_FILE, 'r') as f:
                # do a shorter list to reduce the processing time
                reader = csv.reader(f)
                slist = list(reader)
                stklist = []
                for csvfile in slist[:]:
                    stklist.append(csvfile[0])
                stks = " ".join(stklist)
        except Exception, e:
            stks = ''
        try:
            cmd = "del quotes.csv"
            os.system(cmd)
            if len(stks) == 0:
                cmd = "type *.csv >> quotes.txt"
            else:
                cmd = "type {0} >> quotes.txt".format(stks)
            os.system(cmd)
            cmd = "ren quotes.txt quotes.csv"
            os.system(cmd)
            cmd = "copy quotes.csv {0}".format(target).replace('/', '\\')
            if S.DBG_ALL:
                print cmd
            os.system(cmd)
        except Exception, e:
            print "concat2quotes: ", cmd
            print e
        '''
        with open('output_file.txt','w') as wfd:
        for f in ['seg1.txt','seg2.txt','seg3.txt']:
            with open(f,'rb') as fd:
                shutil.copyfileobj(fd, wfd, 1024*1024*10)
        '''


'''
def xls_to_xlsx(*args, **kw):
    """
    open and convert an XLS file to openpyxl.workbook.Workbook
    ----------
    @param args: args for xlrd.open_workbook
    @param kw: kwargs for xlrd.open_workbook
    @return: openpyxl.workbook.Workbook
    """
    book_xls = xlrd.open_workbook(*args, formatting_info=True,
                                  ragged_rows=True, **kw)
#   book_xlsx = openpyxl.workbook.Workbook()
    book_xlsx = openpyxl.Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])
        for crange in sheet_xls.merged_cells:
            rlo, rhi, clo, chi = crange
            sheet_xlsx.merge_cells(start_row=rlo + 1, end_row=rhi,
                                   start_column=clo + 1, end_column=chi,)

        def _get_xlrd_cell_value(cell):
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                datetime_tup = xlrd.xldate_as_tuple(value, 0)
                if datetime_tup[0:3] == (0, 0, 0):   # time format without date
                    value = datetime.time(*datetime_tup[3:])
                else:
                    value = datetime.datetime(*datetime_tup)
            return value

        for row in range(sheet_xls.nrows):
            sheet_xlsx.append((
                _get_xlrd_cell_value(cell)
                for cell in sheet_xls.row_slice(
                    row, end_colx=sheet_xls.row_len(row))
            ))
    return book_xlsx
'''


class cd:
    """Context manager for changing the current working directory"""
    def __init__(self, newPath):
        self.newPath = os.path.expanduser(newPath)

    def __enter__(self):
        self.savedPath = os.getcwd()
        os.chdir(self.newPath)

    def __exit__(self, etype, value, traceback):
        os.chdir(self.savedPath)


def findAny(substr, infile):
    lines = filter(lambda x: substr in x, open(infile))
    try:
        result = lines[0].rstrip().split(',')
    except Exception:
        result = []
    return result


def findBegins(substr, infile, sep=','):
    with open(infile) as f:
        for line in f:
            if line.startswith(substr + sep):
                result = line.rstrip().split(sep)
                code = result[1]
                return code
    return ''


def getStockCode(shortname, klse_file="scrapers/i3investor/klse.txt", klsemap=''):
    stock_code = findBegins(shortname, klse_file)
    if len(stock_code) == 0 and len(klsemap) > 0:
        # shortname = findBegins(shortname, klsemap, "=")
        shortname = klsemap[shortname]
        if len(shortname) > 0:
            stock_code = findBegins(shortname, klse_file)
    return stock_code


def getStockShortNameById(stock_id, idmap="klse.idmap"):
    if stock_id is None or len(stock_id) <= 0:
        return ''

    found = findAny(stock_id, idmap)
    if len(found) > 0:
        idmap = found[0].split('=')
        return idmap[0]
    return ''


def purgeOldFiles(fltype='*.tgz', days=7):
    current_time = time.time()

    for f in fnmatch.filter(os.listdir('.'), fltype):
        creation_time = os.path.getctime(f)
        count = (current_time - creation_time) // (24 * 3600)
        # print f + ': ' + str(int(count))
        if count >= days:
            os.unlink(f)
            print('{} removed'.format(f))


def mergefiles(directory, fname):
    with cd(directory):
        files = glob.glob(fname + ".*")
        concat = ''.join(open(f).read() for f in files)
        with open(fname, 'w') as f:
            f.write(concat)
        data = file(fname).readlines()
        data.sort()
        with open(fname, 'w') as f:
            for i in range(len(data)):
                f.write(data[i])


'''
def loadDashOptions():
    try:
        with open(S.WORK_DIR + S.MARKET_FILE, 'r') as f:
            reader = csv.reader(f)
            slist = list(reader)
            stklist = []
            for csvfile in slist[:]:
                stklist.append(csvfile[0])
            stks = " ".join(stklist)
    except Exception, e:
        pass

    return slist
'''


def jsonLastDate(counter, datadir=S.DATA_DIR):
    jname = datadir + "json/" + counter + "*.json"
    cmd = "ls -l %s | tail -1 | awk '{print $NF}' | awk -F'[_.]' '{print $2}'" % jname
    lastdate = execshell(cmd)
    return lastdate


if __name__ == '__main__':
    '''
    print wc_line_count("nosuchfile.test")
    print findAny('AIRASIA', '../scrapers/i3investor/klse.txt')
    print getStockCode('AIRASIA', '../scrapers/i3investor/klse.txt')
    print getStockShortNameById('41661', '../scrapers/investingcom/klse.idmap')
    print grepN('/z/data/mpv/KLSE.csv', "2018-10-30")
    print execshell("C:/git/klseScrapers/scripts/lastdate.sh -c danco -d Z:/data")
    cmd = "ls -l /z/data/json/*.json | tail -1 | awk '{print $NF}' | awk -F'[_.]' '{print $2}'"
    print execshell(cmd)
    '''
    print jsonLastDate("DANCO", "/z/data/")
